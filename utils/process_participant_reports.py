import json
import pandas as pd
import argparse
from pathlib import Path
import os
from typing import Dict, Any
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from secretcodes import OPENAI_API_KEY

# progress bar fallback
try:
    from tqdm import tqdm
except Exception:
    def tqdm(iterable, **kwargs):
        return iterable

# openai imports with fallback
try:
    import openai
except Exception:
    openai = None

try:
    from openai import OpenAI
    _HAS_NEW_OPENAI = True
except Exception:
    OpenAI = None
    _HAS_NEW_OPENAI = False

class ParticipantReportProcessor:
    def __init__(self, cases_json_path, image_links_json_path=None, ground_truth_csv_path=None):
        self.cases_json_path = Path(cases_json_path)
        self.dataset_data = self._load_cases_json()
        self.model_name = 'o3'

        if image_links_json_path:
            with open(image_links_json_path, 'r') as f:
                self.image_links = json.load(f)
        else:
            self.image_links = {}

        # load ground truth if provided
        self.ground_truth: Dict[str, Dict[str, Any]] = {}
        if ground_truth_csv_path and Path(ground_truth_csv_path).exists():
            try:
                gt_df = pd.read_csv(ground_truth_csv_path)
                cols = {c.lower().strip(): c for c in gt_df.columns}
                for _, row in gt_df.iterrows():
                    case_id = str(row[cols.get('case')]) if 'case' in cols else None
                    if not case_id:
                        continue
                    self.ground_truth[case_id] = {
                        'Age': row.get(cols.get('age')) if 'age' in cols else None,
                        'Indication': row.get(cols.get('indication')) if 'indication' in cols else None,
                        'Findings': row.get(cols.get('ground truth')) if 'ground truth' in cols else row.get(cols.get('findings'))
                    }
            except Exception as e:
                print(f"[WARN] Could not load ground truth CSV: {e}")

        self._score_cache: Dict[str, Dict[str, Any]] = {}

        self._client = None
        api_key = OPENAI_API_KEY
        if _HAS_NEW_OPENAI and api_key:
            try:
                self._client = OpenAI(api_key=api_key)
            except Exception as e:
                print(f"[WARN] Failed to init OpenAI client: {e}")
        elif _HAS_NEW_OPENAI and not api_key:
            print("[WARN] OPENAI_API_KEY not set. Falling back to heuristic scoring only.")
    
    def _load_cases_json(self):
        if not self.cases_json_path.exists():
            raise FileNotFoundError(f"Cases JSON not found: {self.cases_json_path}")
        with open(self.cases_json_path, 'r') as f:
            data = json.load(f)
        out = {}
        for k, v in data.items():
            sid = v.get('StudyInstanceUid') or k
            out[sid] = v
        return out
    def process_participant_json(self, participant_json_path):
        with open(participant_json_path, 'r') as f:
            participant_data = json.load(f)
        participant_code = participant_data['code_summary']['code']
        
        report_logs = participant_data['report_test_case_logs']
        
        processed_cases = []
        
        for case in report_logs:
            sample_id = case['sample_id']
            dataset_entry = self.ground_truth.get(sample_id) or self.dataset_data.get(sample_id, {})
            if not dataset_entry:
                print(f"[WARN] Missing dataset entry for case {sample_id}")
                dataset_entry = {'PatientAge': None, 'Indication': None, 'Findings': ''}

            processed_case = {
                'Case': sample_id,
                'Age': dataset_entry.get('PatientAge') or dataset_entry.get('Age'),
                'Indication': dataset_entry.get('Indication'),
                'Ground Truth': dataset_entry.get('Findings', ''),
                'Test Type': case['test_type'],
                'Report Text': case['report_text'],
            }

            processed_cases.append(processed_case)
        
        return participant_code, processed_cases
    
    def _make_cache_key(self, case_id: str, text: str, truth: str, test_type: str) -> str:
        return f"{case_id}:::{test_type}:::{hash(text)}:::{hash(truth)}"

    def _llm_score(self, case_id: str, candidate: str, truth: str, age: Any, indication: Any, test_type: str) -> Dict[str, Any]:
        if not candidate or not truth:
            return {
                'score': 0.0 if truth else 1.0,
                'errors': {'a': [], 'b': [], 'c': [], 'd': []},
                'matched_findings': [],
                'summary': 'Empty candidate or ground truth.'
            }
        cache_key = self._make_cache_key(case_id, candidate, truth, test_type)
        if cache_key in self._score_cache:
            return self._score_cache[cache_key]
        age_str = str(age) if age not in (None, '', 'nan') else 'Unknown'
        indication_str = indication if indication else 'None provided'
        reference = f"Findings: {truth}"
        hypothesis = f"Findings: {candidate}"
        prompt = (f'''
                Objective:

                Evaluate the accuracy of a candidate radiology report in comparison to a reference
                radiology report composed by expert radiologists. Only include positive findings, not normal findings. 
                Do not include notes unrelated to clinical findings. 
                
                Process Overview:
                You will be presented with:
                1. The criteria for making a judgment.
                2. The reference radiology report.
                3. The candidate radiology report.
                4. The desired format for your assessment.
                
                1. Criteria for Judgment:
                For each candidate report, determine only the clinically significant errors.

                Errors can fall into one of these categories:
                    a) False report of a finding in the candidate.
                    b) Missing a finding present in the reference.
                    c) Misidentification of a finding's anatomic location/position.
                    d) Misassessment of the severity of a finding.

                Note: Concentrate on the clinical findings rather than the report's writing style.
                Evaluate only the findings that appear in both reports. 
                
                Patient Context:
                    Age: {age_str}
                    Indication: {indication_str}

                IMPORTANT NOTES: 
                    - Evaluate only positive findings, not normal findings. If a finding is normal, it should not be counted in the error or matched finding.
                    - Ignore all references to prior findings and studies. DO NOT COUNT THEM AS ERRORS.
                    - Do NOT penalize the candidate report for omitting specific numeric measurements (e.g., size or dimensions of a nodule/lesion) if the underlying finding is correctly identified. Missing measurements alone is fine since the user writing the candidate report can't measure. They should only be penalized for missing the finding itself.
                    - Do NOT penalize omission of age-appropriate findings that are NOT clinically significant in the context of the indication and patient age.
                    - Do NOT hallucinate or infer findings absent from both reports.
             
                2. Reference Report:
                {reference}
           
                3. Candidate Report:
                {hypothesis}
            
                4. Reporting Your Assessment:
                Format your output as a JSON. Follow this specific format for your output, even if no errors are found:
                ```
                {{
                    "Explanation": "<Explanation>",
                    "ClinicallySignificantErrors": {{
                        "a": ["<Error 1>", "<Error 2>", "...", "<Error n>"],
                        "b": ["<Error 1>", "<Error 2>", "...", "<Error n>"],
                        "c": ["<Error 1>", "<Error 2>", "...", "<Error n>"],
                        "d": ["<Error 1>", "<Error 2>", "...", "<Error n>"]
                    }},
                    "MatchedFindings": ["<Finding 1>", "<Finding 2>", "...", "<Finding n>"]
                }}
                '''
        )
        # Newer client interface (mirrors app usage but simplified) – force json output
        resp = self._client.chat.completions.create(
            model=self.model_name,
            messages=[
                {"role": "system", "content": "You are a precise clinical grading model (CRIMSON)."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"}
        )
        content = resp.choices[0].message.content
        data = json.loads(content)
        raw_errors = data.get('ClinicallySignificantErrors') or {}
        errors = {
            'a': list(raw_errors.get('a', []) or []),
            'b': list(raw_errors.get('b', []) or []),
            'c': list(raw_errors.get('c', []) or []),
            'd': list(raw_errors.get('d', []) or []),
        }
        matched = list(data.get('MatchedFindings') or [])
        total_matched = len(matched)
        total_errors = sum(len(errors[k]) for k in ('a','b','c','d'))
        score = 1.0 if total_errors == 0 else total_matched / (total_matched + total_errors)
        result = {
            'score': score,
            'errors': errors,
            'matched_findings': matched,
            'summary': data.get('Explanation') or ''
        }
        
        self._score_cache[cache_key] = result
        return result
    
    def create_excel_report(self, participant_json_paths, output_path):
        """Create Excel file with participant data."""
        if isinstance(participant_json_paths, str):
            participant_json_paths = [participant_json_paths]

        all_data = {}

        for participant_json_path in tqdm(participant_json_paths, desc='Participants'):
            participant_code, processed_cases = self.process_participant_json(participant_json_path)
            if not processed_cases:
                print(f"No valid cases found for participant {participant_code}")
                continue

            df = pd.DataFrame(processed_cases)
            pre_cases = df[df['Test Type'] == 'pre'].copy()
            post_cases = df[df['Test Type'] == 'post'].copy()
            final_data = []
            case_indices = sorted(df['Case'].unique())

            for case_idx in tqdm(case_indices, desc=f'Cases {participant_code}', leave=False):
                case_pre = pre_cases[pre_cases['Case'] == case_idx]
                case_post = post_cases[post_cases['Case'] == case_idx]
                if case_pre.empty and case_post.empty:
                    continue
                ref_row = case_pre.iloc[0] if not case_pre.empty else case_post.iloc[0]
                ground_truth = ref_row['Ground Truth'] or ''
                age = ref_row['Age']
                indication = ref_row['Indication']
                pre_text = case_pre.iloc[0]['Report Text'] if not case_pre.empty else ''
                post_text = case_post.iloc[0]['Report Text'] if not case_post.empty else ''

                pre_score_payload = self._llm_score(case_idx, pre_text, ground_truth, age, indication, 'pre') if pre_text else {
                    'score': None, 'errors': {'a': [], 'b': [], 'c': [], 'd': []}, 'matched_findings': [], 'summary': ''
                }
                post_score_payload = self._llm_score(case_idx, post_text, ground_truth, age, indication, 'post') if post_text else {
                    'score': None, 'errors': {'a': [], 'b': [], 'c': [], 'd': []}, 'matched_findings': [], 'summary': ''
                }

                # Map error categories to descriptive names
                error_labels = {
                    'a': 'False Finding',          # False report of a finding
                    'b': 'Missing Finding',        # Missing a finding present in reference
                    'c': 'Location Error',         # Misidentification of location
                    'd': 'Severity Error'          # Misassessment of severity
                }

                def _enum(items):
                    if not items:
                        return ''
                    # Numbered list each on new line: 1) item
                    return '\n'.join(f"{i}. {v}" for i, v in enumerate(items, 1))

                final_row = {
                    'Case': case_idx,
                    'Image Link': self.image_links.get(case_idx, ''),
                    'Age': age,
                    'Indication': indication,
                    'Ground Truth': ground_truth,
                    'Report_pre': pre_text,
                    'Score_pre': pre_score_payload['score'],
                    'Matched Findings (pre)': _enum(pre_score_payload['matched_findings']),
                    'Ground Truth (repeat)': ground_truth,  # duplicated immediately after pre matched findings
                    'Report_post': post_text,
                    'Score_post': post_score_payload['score'],
                    'Matched Findings (post)': _enum(post_score_payload['matched_findings']),
                }

                # Add descriptive error columns (prefixed with pre/post)
                for cat, label in error_labels.items():
                    pre_errs = pre_score_payload['errors'].get(cat, [])
                    post_errs = post_score_payload['errors'].get(cat, [])
                    final_row[f"pre {label}"] = _enum(pre_errs)
                    final_row[f"post {label}"] = _enum(post_errs)

                final_data.append(final_row)

            all_data[participant_code] = pd.DataFrame(final_data)
            print(f"Processed {participant_json_path} -> {participant_code} ({len(final_data)} cases)")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for participant_code, df in all_data.items():
                # Explicit column ordering to preserve original intended sequence
                # User-requested primary ordering (exact sequence from final_row construction)
                # Assemble desired order: pre core block + pre errors, then post core + post errors
                pre_core = [
                    'Case', 'Image Link', 'Age', 'Indication', 'Ground Truth',
                    'Report_pre', 'Score_pre', 'Matched Findings (pre)'
                ]
                pre_errors = ['pre False Finding', 'pre Missing Finding', 'pre Location Error', 'pre Severity Error']
                post_core = ['Ground Truth (repeat)', 'Report_post', 'Score_post', 'Matched Findings (post)']
                post_errors = ['post False Finding', 'post Missing Finding', 'post Location Error', 'post Severity Error']
                desired = [c for c in pre_core + pre_errors + post_core + post_errors if c in df.columns]
                leftover = [c for c in df.columns if c not in desired]
                df = df[desired + leftover]
                df.to_excel(writer, sheet_name=participant_code, index=False)
                worksheet = writer.sheets[participant_code]

                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=2)
                    image_link = cell.value
                    cell.hyperlink = image_link
                    cell.value = 'View Image'
                    cell.style = 'Hyperlink'

                wrap_cols = {c for c in df.columns if any(k in c.lower() for k in ['report', 'ground truth', 'error', 'matched', 'finding'])}
                for col_idx, col_name in enumerate(df.columns, start=1):
                    try:
                        values = df[col_name].astype(str).tolist()
                    except Exception:
                        values = []
                    max_len = max([len(col_name)] + [len(v) for v in values]) if values else len(col_name)
                    max_len = min(max_len, 80)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)
                    if col_name in wrap_cols:
                        for row_num in range(1, len(df) + 2):
                            c = worksheet.cell(row=row_num, column=col_idx)
                            c.alignment = Alignment(wrap_text=True, vertical='top')

        print(f"Excel report created: {output_path}")
        print(f"Total participants: {len(all_data)}")
        return all_data
    

def main():
    parser = argparse.ArgumentParser(description='Process participant report JSON files and generate Excel reports')
    
    parser.add_argument('--json-dir', 
                       help='Directory containing participant JSON files')
    
    parser.add_argument('-o', '--output', 
                       help='Output Excel file path')
    
    parser.add_argument('-c', '--cases-json', 
                       default='/home/baharoon/RadGame/data/report_test_cases.json',
                       help='Path to report test cases JSON (single file)')

    parser.add_argument('-l', '--image-links', 
                       help='Path to JSON file containing case_id to image link mapping')
    parser.add_argument('-g', '--ground-truth-csv',
                       help='Path to ground truth CSV (Case,Age,Indication,Ground Truth)')
    
    args = parser.parse_args()
  
    processor = ParticipantReportProcessor(
        args.cases_json,
        args.image_links,
        ground_truth_csv_path=args.ground_truth_csv
    )
    
    json_dir = Path(args.json_dir)
    json_files = list(json_dir.glob("*.json"))
    
    if not json_files:
        print(f"No JSON files found in directory: {args.json_dir}")
        return
    
    print(f"Found {len(json_files)} JSON files in {args.json_dir}")
    
    processor.create_excel_report(json_files, args.output)
    print(f"\nSuccess! Generated Excel report: {args.output}")


if __name__ == "__main__":
    main()
